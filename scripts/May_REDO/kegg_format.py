""" Purpose of this script is to just format a string from KEGG to make sure
    just have a list of gene names, this is in response to Laura's request
    that this was an interesting gene set.
"""

################################################################################
                    # Environment setup #
################################################################################
import os, sys
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scripts.May_REDO.helpers as hs
import scripts.May_REDO.DE_figure_helpers as dhs
from sklearn.preprocessing import scale

out_plots = 'figure_components/PseudoLimma/DE_figures/'

hs.setUp()

################################################################################
                    # Loading & formatting the file #
################################################################################
data_dir = 'data/DE_out/Pseudo_TMM_Limma_Voom/data/'
lcpms = pd.read_csv(data_dir+'lcpms.txt', sep='\t')
human_genes = [gene.replace('hg38-', '') for gene in lcpms.index
               if 'hg38-' in gene]
ecm_recepts = []
kegg_file = open('scripts/May_REDO/ECM-receptor_KEGG.txt', 'r')
n_lines = 0
for line in kegg_file:
    gene_info = line.split('|')[-1]
    ref_genes = gene_info.split(';')[0].replace(' (RefSeq) ', '')\
                    .strip('\n').split(', ')
    cnt = 0
    for gene in ref_genes:
        if gene in human_genes:
            ecm_recepts.append( gene )
            cnt+=1
        if cnt>1:
            print(ref_genes)
    if cnt == 0 :
        print(ref_genes)
    n_lines += 1

# Adding this into the excel sheet for giotto spot-wise enrichment:
# data/third_party_data/Barnes2018_NatureCellBiology_MesynchymalGlioblastoma.xlsx
for gene in ecm_recepts:
    print(gene)

import os
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

ecm_df = pd.DataFrame(ecm_recepts, columns=['Kegg_ECM-Receptor'])

append_df_to_excel('data/third_party_data/Barnes2018_NatureCellBiology_MesynchymalGlioblastoma.xlsx',
                   ecm_df, sheet_name='AddByBrad',
                   engine='openpyxl', index=False)









